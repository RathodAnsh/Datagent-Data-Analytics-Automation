#5. excel_dashboard.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    BarChart, 
    LineChart, 
    PieChart, 
    AreaChart, 
    ScatterChart, 
    Reference
)
from openpyxl.styles import Font, Alignment, PatternFill

def apply_derivation(df, task):
    """ 
    Executes dynamic Python logic for Excel. 
    Handles multi-line logic like date conversions and math.
    """
    df_copy = df.copy()
    derivation = task.get('derivation')
    
    # Check if derivation exists and logic is not empty/None
    if derivation and derivation.get('python_logic'):
        try:
            # Use the State Dict logic to ensure changes to 'df' stick
            state = {"pd": pd, "df": df_copy}
            logic = derivation['python_logic']
            parts = [p.strip() for p in logic.split(';') if p.strip()]
            
            for part in parts:
                exec(part, state, state)
            
            # Return the updated dataframe from the state dictionary
            return state["df"] 
        except Exception as e:
            print(f"Excel Logic Error: {e}")
            return df_copy
    return df_copy

def build_excel_dashboard(df, plan):
    wb = Workbook()
    
    # 1. RAW DATA SHEET
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_raw.append(r)
    
    for cell in ws_raw[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # 2. ANALYSIS SHEETS
    for i, task in enumerate(plan):
        try:
            # --- APPLY DYNAMIC LOGIC ---
            processed_df = apply_derivation(df, task)

            # --- FIX: Ensure aggregation is lowercase (e.g., 'sum' instead of 'Sum') ---
            agg_func = (task.get('aggregation') or 'sum').lower()

            # Perform aggregation on the processed dataframe
            chart_data = processed_df.groupby(task["x_axis"])[task["y_axis"]].agg(agg_func).reset_index()

            sheet_name = f"Analysis {i+1}"
            ws = wb.create_sheet(title=sheet_name)
            
            # --- TABLE SECTION ---
            ws.merge_cells("A1:C1")
            ws["A1"] = f"Question: {task['question']}"
            ws["A1"].font = Font(bold=True, size=11)
            ws["A1"].alignment = Alignment(horizontal="left", wrap_text=True)
            ws.row_dimensions[1].height = 30 
            
            # Write aggregated data starting at Row 3
            for r_idx, row in enumerate(dataframe_to_rows(chart_data, index=False, header=True), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3: # Header styling
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # --- CHART SECTION ---
            c_type = (task.get('chart_type') or 'bar').lower()
            
            if "line" in c_type:
                chart = LineChart()
            elif "pie" in c_type:
                chart = PieChart()
                if "doughnut" in c_type: chart.type = "doughnut"
            elif "area" in c_type:
                chart = AreaChart()
            elif "scatter" in c_type:
                chart = ScatterChart()
            else:
                chart = BarChart()
                chart.type = "col"

            chart.title = f"{task['y_axis']} by {task['x_axis']}"
            chart.style = 10 
            
            # --- FIX: Precise Reference Mapping ---
            # Data starts at Row 3 (Headers). Row 4 is first data row.
            data_rows = len(chart_data)
            
            # Values: Column 2 (Y-axis), Rows 3 to (3 + rows)
            values = Reference(ws, min_col=2, min_row=3, max_row=3 + data_rows)
            # Categories: Column 1 (X-axis), Rows 4 to (3 + rows)
            categories = Reference(ws, min_col=1, min_row=4, max_row=3 + data_rows)
            
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, "E2") # Moved to E2 to give the table more room
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20

        except Exception as e:
            print(f"Error generating sheet {i+1}: {e}")
            continue

    file_path = "automated_analysis_report.xlsx"
    wb.save(file_path)
    return file_path